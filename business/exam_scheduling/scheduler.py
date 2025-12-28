# exam_optimizer.py
# Final Exam Scheduler v1.1 - OR-Tools CP-SAT
# Author branding lives in GUI

from __future__ import annotations

import re
from dataclasses import dataclass
from typing import Dict, List, Tuple, Any, Optional

# pandas is optional at import-time (GUI shows friendly install hint)
try:
    import pandas as pd
except Exception:
    pd = None


# ----------------------------- Helpers -----------------------------

def require_pandas():
    if pd is None:
        raise ImportError(
            "pandas is not installed.\n\n"
            "Install with:\n"
            "python -m pip install pandas openpyxl ortools"
        )


def normalize_program(x: Any) -> str:
    return str(x).strip().upper()


def normalize_str(x: Any) -> str:
    if x is None:
        return ""
    s = str(x).strip()
    return "" if s.lower() in ("nan", "none") else s


def split_courses(cell: Any) -> List[str]:
    if cell is None:
        return []
    s = str(cell)
    if s.lower() in ("nan", "none"):
        return []
    parts = [p.strip() for p in s.split(",")]
    return [p for p in parts if p and p.lower() not in ("nan", "none")]


def normalize_date_ignore_year(x: Any):
    """
    Parse date then force year=2000 so comparisons depend on month/day only.
    Returns pandas Timestamp or NaT.
    """
    require_pandas()
    ts = pd.to_datetime(x, errors="coerce")
    if pd.isna(ts):
        return pd.NaT
    return pd.Timestamp(year=2000, month=int(ts.month), day=int(ts.day))


def time_to_min(x: Any) -> int:
    """
    Accept '09:30' or 930 / 1130 style. Return minutes from 00:00.
    """
    if x is None:
        return 0
    s = str(x).strip()
    if not s or s.lower() in ("nan", "none"):
        return 0

    if ":" in s:
        hh, mm = s.split(":")
        return int(hh) * 60 + int(mm)

    s2 = re.sub(r"\D", "", s)
    if not s2:
        return 0
    n = int(s2)
    h, m = n // 100, n % 100
    return h * 60 + m


def fmt_hhmm(minutes: int) -> str:
    h = minutes // 60
    m = minutes % 60
    return f"{h:02d}:{m:02d}"


def mmdd_str(ts) -> str:
    return ts.strftime("%m-%d")


def safe_sheet_name(name: str) -> str:
    name = re.sub(r"[\[\]\:\*\?\/\\]", "_", name)
    return name[:31]


# ----------------------------- Diagnostics -----------------------------

@dataclass
class DiagnosticsResult:
    diag: Dict[str, Any]
    dfs: Dict[str, "pd.DataFrame"]


# ----------------------------- Loaders -----------------------------

def _load_inputs(
    regs_path: str,
    courses_master_path: str,
    calendar_path: str,
    slot_capacity_path: str,
    constraints_path: str,
):
    require_pandas()

    regs_df = pd.read_excel(regs_path, sheet_name="Regs")
    courses_df = pd.read_excel(courses_master_path, sheet_name="Courses")
    cal_df = pd.read_excel(calendar_path, sheet_name="Calendar")
    cap_df = pd.read_excel(slot_capacity_path, sheet_name="SlotCapacity")
    fixed_df = pd.read_excel(constraints_path, sheet_name="FixedAssignments")
    try:
        balance_df = pd.read_excel(constraints_path, sheet_name="BalanceSettings")
    except Exception:
        balance_df = pd.DataFrame()

    # Required columns checks
    req_regs = {"ID", "Program", "COURSES"}
    req_courses = {"CourseCode", "Program", "ExamGroup"}
    req_cal = {"Date", "SlotID", "Start", "End"}
    req_cap = {"Date", "SlotID", "CapacityStudents"}

    if not req_regs.issubset(regs_df.columns):
        raise ValueError(f"regs.xlsx (Regs) must include columns: {sorted(req_regs)}")
    if not req_courses.issubset(courses_df.columns):
        raise ValueError(f"courses_master.xlsx (Courses) must include columns: {sorted(req_courses)}")
    if not req_cal.issubset(cal_df.columns):
        raise ValueError(f"exam_calendar.xlsx (Calendar) must include columns: {sorted(req_cal)}")
    if not req_cap.issubset(cap_df.columns):
        raise ValueError(f"slot_capacity.xlsx (SlotCapacity) must include columns: {sorted(req_cap)}")

    regs_df = regs_df.copy()
    courses_df = courses_df.copy()
    cal_df = cal_df.copy()
    cap_df = cap_df.copy()
    fixed_df = fixed_df.copy()
    balance_df = balance_df.copy()

    # regs
    regs_df["ID"] = regs_df["ID"].astype(str).str.strip()
    regs_df["Program"] = regs_df["Program"].apply(normalize_program)
    if "NAME" in regs_df.columns:
        regs_df["NAME"] = regs_df["NAME"].apply(normalize_str)

    # courses
    courses_df["CourseCode"] = courses_df["CourseCode"].astype(str).str.strip()
    courses_df["Program"] = courses_df["Program"].apply(normalize_program)
    courses_df["ExamGroup"] = courses_df["ExamGroup"].apply(normalize_str)
    if "CourseName" not in courses_df.columns:
        courses_df["CourseName"] = ""
    courses_df["CourseName"] = courses_df["CourseName"].apply(normalize_str)

    if "DurationMin" not in courses_df.columns:
        courses_df["DurationMin"] = 120
    courses_df["DurationMin"] = pd.to_numeric(courses_df["DurationMin"], errors="coerce").fillna(120).astype(int)

    # calendar
    cal_df["DateN"] = cal_df["Date"].apply(normalize_date_ignore_year)
    if cal_df["DateN"].isna().any():
        raise ValueError("Some Calendar Date values could not be parsed. Fix exam_calendar.xlsx.")

    cal_df["StartMin"] = cal_df["Start"].apply(time_to_min)
    cal_df["EndMin"] = cal_df["End"].apply(time_to_min)
    cal_df["SlotDurationMin"] = (cal_df["EndMin"] - cal_df["StartMin"]).astype(int)
    if (cal_df["SlotDurationMin"] <= 0).any():
        raise ValueError("Calendar has invalid Start/End: End must be after Start for every slot.")
    cal_df["SlotKey"] = cal_df["DateN"].astype(str) + " | " + cal_df["SlotID"].astype(str)

    # capacity
    cap_df["DateN"] = cap_df["Date"].apply(normalize_date_ignore_year)
    if cap_df["DateN"].isna().any():
        raise ValueError("Some SlotCapacity Date values could not be parsed. Fix slot_capacity.xlsx.")
    cap_df["SlotKey"] = cap_df["DateN"].astype(str) + " | " + cap_df["SlotID"].astype(str)
    cap_df["CapacityStudents"] = pd.to_numeric(cap_df["CapacityStudents"], errors="coerce").fillna(0).astype(int)

    # fixed
    if fixed_df is None or fixed_df.empty:
        fixed_df = pd.DataFrame(columns=["ExamGroup", "Date", "SlotID", "DateN", "SlotKey"])
    else:
        req_fixed = {"ExamGroup", "Date", "SlotID"}
        if not req_fixed.issubset(set(fixed_df.columns)):
            raise ValueError(f"constraints.xlsx (FixedAssignments) must include columns: {sorted(req_fixed)}")
        fixed_df["ExamGroup"] = fixed_df["ExamGroup"].apply(normalize_str)
        fixed_df["DateN"] = fixed_df["Date"].apply(normalize_date_ignore_year)
        fixed_df["SlotKey"] = fixed_df["DateN"].astype(str) + " | " + fixed_df["SlotID"].astype(str)

    return regs_df, courses_df, cal_df, cap_df, fixed_df, balance_df


def _build_enrollments(regs_df, courses_df) -> Tuple["pd.DataFrame", "pd.DataFrame"]:
    """
    Expand regs -> enrollments and join with courses_master using rule:
    join (CourseCode, Program) first; fallback to (CourseCode, Program=ALL).
    Returns (enroll_df, missing_df)
    """
    require_pandas()

    rows = []
    for _, r in regs_df.iterrows():
        sid = str(r["ID"]).strip()
        prog = normalize_program(r["Program"])
        for cc in split_courses(r["COURSES"]):
            rows.append({"StudentID": sid, "Program": prog, "CourseCode": str(cc).strip()})

    enroll = pd.DataFrame(rows)
    if enroll.empty:
        raise ValueError("No enrollments parsed from regs.xlsx (COURSES might be empty).")

    # Join Program-specific first
    courses_exact = courses_df[courses_df["Program"] != "ALL"][
        ["CourseCode", "Program", "CourseName", "ExamGroup", "DurationMin"]
    ]
    courses_all = courses_df[courses_df["Program"] == "ALL"][
        ["CourseCode", "CourseName", "ExamGroup", "DurationMin"]
    ]

    enroll = enroll.merge(courses_exact, on=["CourseCode", "Program"], how="left")

    # fallback to ALL
    missing_mask = enroll["ExamGroup"].isna() | (enroll["ExamGroup"].astype(str).str.strip() == "")
    if missing_mask.any():
        fallback = enroll.loc[missing_mask, ["CourseCode"]].merge(courses_all, on="CourseCode", how="left")
        enroll.loc[missing_mask, "CourseName"] = fallback["CourseName"].values
        enroll.loc[missing_mask, "ExamGroup"] = fallback["ExamGroup"].values
        enroll.loc[missing_mask, "DurationMin"] = fallback["DurationMin"].values

    still_missing = enroll["ExamGroup"].isna() | (enroll["ExamGroup"].astype(str).str.strip() == "")
    if still_missing.any():
        miss = enroll.loc[still_missing, ["CourseCode"]].drop_duplicates().sort_values("CourseCode")
        missing_df = miss.rename(columns={"CourseCode": "MissingCourseCode"})
    else:
        missing_df = pd.DataFrame(columns=["MissingCourseCode"])

    enroll["CourseName"] = enroll["CourseName"].fillna("").astype(str)
    enroll["ExamGroup"] = enroll["ExamGroup"].fillna("").astype(str)
    enroll["DurationMin"] = pd.to_numeric(enroll["DurationMin"], errors="coerce").fillna(120).astype(int)

    return enroll, missing_df


# ----------------------------- Diagnostics -----------------------------

def _compute_diagnostics(
    regs_df, courses_df, cal_df, cap_df, fixed_df, enroll_df, missing_df
) -> DiagnosticsResult:
    require_pandas()

    n_students = enroll_df["StudentID"].nunique()
    n_programs = enroll_df["Program"].nunique()
    n_enroll = len(enroll_df)

    regs_codes = set(enroll_df["CourseCode"].unique().tolist())
    master_codes = set(courses_df["CourseCode"].unique().tolist())
    found_codes = regs_codes.intersection(master_codes)

    examgroups = sorted([g for g in enroll_df["ExamGroup"].dropna().unique().tolist() if str(g).strip() != ""])
    n_examgroups = len(examgroups)

    slots = cal_df[["SlotKey", "DateN", "SlotID", "StartMin", "EndMin", "SlotDurationMin"]].copy()
    slot_keys = slots["SlotKey"].tolist()
    n_slots = len(slot_keys)
    unique_days = sorted(slots["DateN"].unique())
    n_days = len(unique_days)

    # SlotStats
    slot_stats = slots.merge(cap_df[["SlotKey", "CapacityStudents"]], on="SlotKey", how="left")
    slot_stats["CapacityStudents"] = slot_stats["CapacityStudents"].fillna(10**9).astype(int)
    slot_stats["Date_MMDD"] = slot_stats["DateN"].apply(mmdd_str)
    slot_stats["Start"] = slot_stats["StartMin"].apply(fmt_hhmm)
    slot_stats["End"] = slot_stats["EndMin"].apply(fmt_hhmm)
    slot_stats = slot_stats[["Date_MMDD", "SlotID", "Start", "End", "SlotDurationMin", "CapacityStudents", "SlotKey"]].copy()

    # ExamGroup stats
    g_students = enroll_df.groupby("ExamGroup")["StudentID"].nunique().to_dict()
    g_duration = enroll_df.groupby("ExamGroup")["DurationMin"].max().to_dict()
    slot_dur_map = dict(zip(slots["SlotKey"], slots["SlotDurationMin"]))

    eg_rows = []
    for g in examgroups:
        dur = int(g_duration.get(g, 120))
        feasible = sum(1 for sk in slot_keys if int(slot_dur_map.get(sk, 0)) >= dur)
        eg_rows.append({
            "ExamGroup": g,
            "Students": int(g_students.get(g, 0)),
            "DurationMin": dur,
            "FeasibleSlotsByDuration": int(feasible),
        })
    eg_stats = pd.DataFrame(eg_rows).sort_values(["FeasibleSlotsByDuration", "Students"], ascending=[True, False])

    # conflict density
    student_groups = enroll_df.groupby("StudentID")["ExamGroup"].apply(lambda s: list(sorted(set([x for x in s if str(x).strip() != ""])))).to_dict()
    max_exams_student = max((len(v) for v in student_groups.values()), default=0)
    total_pairs = sum((len(v) * (len(v) - 1)) // 2 for v in student_groups.values())

    # Fixed assignment issues
    fixed_issues_rows = []
    slotkey_set = set(slot_keys)
    eg_set = set(examgroups)
    if fixed_df is not None and not fixed_df.empty:
        for _, r in fixed_df.iterrows():
            eg = normalize_str(r.get("ExamGroup"))
            sk = normalize_str(r.get("SlotKey"))
            if not eg and not sk:
                continue
            issue = None
            if eg not in eg_set:
                issue = "Unknown ExamGroup (not present in regs enrollments)"
            elif sk not in slotkey_set:
                issue = "SlotKey not found in Calendar"
            if issue:
                fixed_issues_rows.append({"ExamGroup": eg, "SlotKey": sk, "Issue": issue})
    fixed_issues_df = pd.DataFrame(fixed_issues_rows)

    diag = {
        "students": int(n_students),
        "programs": int(n_programs),
        "enrollments": int(n_enroll),
        "unique_coursecodes_in_regs": int(len(regs_codes)),
        "coursecodes_found_in_master": int(len(found_codes)),
        "missing_coursecodes_count": int(len(missing_df)) if missing_df is not None and not missing_df.empty else 0,
        "missing_coursecodes_list": missing_df["MissingCourseCode"].tolist() if missing_df is not None and not missing_df.empty else [],
        "examgroups": int(n_examgroups),
        "slots": int(n_slots),
        "unique_days": int(n_days),
        "max_exams_per_student": int(max_exams_student),
        "total_pairwise_exam_pairs": int(total_pairs),
        "fixed_assignment_issues": int(len(fixed_issues_df)) if fixed_issues_df is not None and not fixed_issues_df.empty else 0,
    }

    dfs = {
        "EnrollmentsPreview": enroll_df.head(500).copy(),
        "MissingCourseCodes": (missing_df.copy() if missing_df is not None else pd.DataFrame()),
        "ExamGroupStats": eg_stats.copy(),
        "FixedAssignmentIssues": fixed_issues_df.copy(),
        "SlotStats": slot_stats.copy(),
    }
    return DiagnosticsResult(diag=diag, dfs=dfs)


def save_diagnostics_excel(dfs: Dict[str, "pd.DataFrame"], output_path: str):
    require_pandas()
    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
        for name, df in dfs.items():
            df.to_excel(writer, index=False, sheet_name=safe_sheet_name(name))


# ----------------------------- Courses Report -----------------------------

def generate_courses_report(regs_path: str, courses_master_path: str) -> Tuple["pd.DataFrame", "pd.DataFrame"]:
    """
    Returns:
      report_df: CourseCode, ResolvedCourseName, ResolvedExamGroup, DurationMin, TotalStudents, + per-program counts
      issues_df: rows where NOT TITLED and/or NOT GROUPED and/or MissingInMaster
    """
    require_pandas()

    regs_df = pd.read_excel(regs_path, sheet_name="Regs")
    courses_df = pd.read_excel(courses_master_path, sheet_name="Courses")

    req_regs = {"ID", "Program", "COURSES"}
    req_courses = {"CourseCode", "Program", "ExamGroup"}
    if not req_regs.issubset(regs_df.columns):
        raise ValueError(f"regs.xlsx (Regs) must include columns: {sorted(req_regs)}")
    if not req_courses.issubset(courses_df.columns):
        raise ValueError(f"courses_master.xlsx (Courses) must include columns: {sorted(req_courses)}")

    regs_df = regs_df.copy()
    courses_df = courses_df.copy()

    regs_df["ID"] = regs_df["ID"].astype(str).str.strip()
    regs_df["Program"] = regs_df["Program"].apply(normalize_program)

    courses_df["CourseCode"] = courses_df["CourseCode"].astype(str).str.strip()
    courses_df["Program"] = courses_df["Program"].apply(normalize_program)
    courses_df["ExamGroup"] = courses_df["ExamGroup"].apply(normalize_str)
    if "CourseName" not in courses_df.columns:
        courses_df["CourseName"] = ""
    courses_df["CourseName"] = courses_df["CourseName"].apply(normalize_str)
    if "DurationMin" not in courses_df.columns:
        courses_df["DurationMin"] = 120
    courses_df["DurationMin"] = pd.to_numeric(courses_df["DurationMin"], errors="coerce").fillna(120).astype(int)

    # enrollments expansion
    rows = []
    for _, r in regs_df.iterrows():
        sid = str(r["ID"]).strip()
        prog = normalize_program(r["Program"])
        for cc in split_courses(r["COURSES"]):
            rows.append({"StudentID": sid, "Program": prog, "CourseCode": str(cc).strip()})
    enroll = pd.DataFrame(rows)
    if enroll.empty:
        raise ValueError("No enrollments parsed from regs.xlsx.")

    # Resolve master fields with same join rule
    courses_exact = courses_df[courses_df["Program"] != "ALL"][["CourseCode", "Program", "CourseName", "ExamGroup", "DurationMin"]]
    courses_all = courses_df[courses_df["Program"] == "ALL"][["CourseCode", "CourseName", "ExamGroup", "DurationMin"]]

    merged = enroll.merge(courses_exact, on=["CourseCode", "Program"], how="left")
    missing_mask = merged["ExamGroup"].isna() | (merged["ExamGroup"].astype(str).str.strip() == "")
    if missing_mask.any():
        fb = merged.loc[missing_mask, ["CourseCode"]].merge(courses_all, on="CourseCode", how="left")
        merged.loc[missing_mask, "CourseName"] = fb["CourseName"].values
        merged.loc[missing_mask, "ExamGroup"] = fb["ExamGroup"].values
        merged.loc[missing_mask, "DurationMin"] = fb["DurationMin"].values

    merged["CourseName"] = merged["CourseName"].fillna("").astype(str)
    merged["ExamGroup"] = merged["ExamGroup"].fillna("").astype(str)
    merged["DurationMin"] = pd.to_numeric(merged["DurationMin"], errors="coerce").fillna(120).astype(int)

    # Counts
    total_counts = merged.groupby("CourseCode")["StudentID"].nunique().rename("TotalStudents").reset_index()
    by_prog = merged.groupby(["CourseCode", "Program"])["StudentID"].nunique().reset_index()
    pivot = by_prog.pivot(index="CourseCode", columns="Program", values="StudentID").fillna(0).astype(int).reset_index()

    # Resolved fields per course (pick the most common non-empty name/group if multiple)
    def mode_nonempty(s):
        s2 = [x for x in s.astype(str).tolist() if str(x).strip() != "" and str(x).lower() not in ("nan", "none")]
        if not s2:
            return ""
        return pd.Series(s2).mode().iloc[0]

    fields = merged.groupby("CourseCode").agg(
        ResolvedCourseName=("CourseName", mode_nonempty),
        ResolvedExamGroup=("ExamGroup", mode_nonempty),
        DurationMin=("DurationMin", "max"),
    ).reset_index()

    report = fields.merge(total_counts, on="CourseCode", how="left").merge(pivot, on="CourseCode", how="left")
    report["TotalStudents"] = report["TotalStudents"].fillna(0).astype(int)

    # Flags
    report["NOT_TITLED"] = report["ResolvedCourseName"].astype(str).str.strip().eq("")
    report["NOT_GROUPED"] = report["ResolvedExamGroup"].astype(str).str.strip().eq("")

    # Missing in master (course code doesn't exist at all)
    master_codes = set(courses_df["CourseCode"].unique().tolist())
    report["MISSING_IN_MASTER"] = ~report["CourseCode"].isin(master_codes)

    # issues dataframe
    issues = report.loc[
        report["NOT_TITLED"] | report["NOT_GROUPED"] | report["MISSING_IN_MASTER"],
        ["CourseCode", "ResolvedCourseName", "ResolvedExamGroup", "DurationMin", "TotalStudents", "NOT_TITLED", "NOT_GROUPED", "MISSING_IN_MASTER"]
    ].copy()

    # Sort nicely
    report = report.sort_values(["TotalStudents", "CourseCode"], ascending=[False, True]).reset_index(drop=True)
    issues = issues.sort_values(["MISSING_IN_MASTER", "NOT_GROUPED", "NOT_TITLED", "TotalStudents"], ascending=[False, False, False, False]).reset_index(drop=True)

    return report, issues


def save_courses_report_excel(report_df: "pd.DataFrame", issues_df: "pd.DataFrame", output_path: str):
    """
    Saves a Courses Report Excel and highlights problematic rows.
    """
    require_pandas()
    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
        report_df.to_excel(writer, index=False, sheet_name="CoursesReport")
        issues_df.to_excel(writer, index=False, sheet_name="Issues")

    # Highlight using openpyxl
    try:
        from openpyxl import load_workbook
        from openpyxl.styles import PatternFill

        wb = load_workbook(output_path)
        ws = wb["CoursesReport"]

        # Find flag columns
        headers = [c.value for c in ws[1]]
        def col_idx(name: str) -> Optional[int]:
            try:
                return headers.index(name) + 1
            except ValueError:
                return None

        c_not_titled = col_idx("NOT_TITLED")
        c_not_grouped = col_idx("NOT_GROUPED")
        c_missing = col_idx("MISSING_IN_MASTER")

        fill_warn = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")  # light yellow
        fill_bad = PatternFill(start_color="F8CBAD", end_color="F8CBAD", fill_type="solid")   # light red

        for r in range(2, ws.max_row + 1):
            nt = bool(ws.cell(r, c_not_titled).value) if c_not_titled else False
            ng = bool(ws.cell(r, c_not_grouped).value) if c_not_grouped else False
            ms = bool(ws.cell(r, c_missing).value) if c_missing else False

            if ms:
                for c in range(1, ws.max_column + 1):
                    ws.cell(r, c).fill = fill_bad
            elif nt or ng:
                for c in range(1, ws.max_column + 1):
                    ws.cell(r, c).fill = fill_warn

        wb.save(output_path)
    except Exception:
        # If highlighting fails, keep the Excel content at least
        pass


# ----------------------------- Main Scheduler -----------------------------

def run_final_exam_scheduler(
    regs_path: str,
    courses_master_path: str,
    calendar_path: str,
    slot_capacity_path: str,
    constraints_path: str,
    output_path: str = "Final_Exam_Schedule.xlsx",
    rest_days: int = 1,
    time_limit_sec: int = 40,
    workers: int = 8,
    diagnostics_only: bool = False,
):
    """
    If diagnostics_only=True:
      - DO NOT import OR-Tools
      - DO NOT build model
      - DO NOT solve
      Return (DiagnosticsDict, DiagnosticsDataFrames)

    If diagnostics_only=False:
      Solve and write output excel.
      Return (master_df, program_sheets, cap_report_df, rest_viol_df, summary_df)
    """
    require_pandas()

    regs_df, courses_df, cal_df, cap_df, fixed_df, balance_df = _load_inputs(
        regs_path, courses_master_path, calendar_path, slot_capacity_path, constraints_path
    )
    enroll_df, missing_df = _build_enrollments(regs_df, courses_df)

    diag_res = _compute_diagnostics(regs_df, courses_df, cal_df, cap_df, fixed_df, enroll_df, missing_df)

    # In diagnostics mode: do NOT raise; just return diagnostics even if missing exists
    if diagnostics_only:
        return diag_res.diag, diag_res.dfs

    # In solve mode: missing course codes is a hard stop
    if missing_df is not None and not missing_df.empty:
        missing_list = missing_df["MissingCourseCode"].tolist()
        raise ValueError(
            "Some CourseCodes in regs.xlsx are missing from courses_master.xlsx.\n\n"
            f"Missing ({len(missing_list)}): {', '.join(missing_list[:50])}"
            + (" ..." if len(missing_list) > 50 else "")
            + "\n\nFix:\n"
              "- Add these CourseCodes to courses_master.xlsx (sheet Courses)\n"
              "- Either with Program=ALL or specific Program rows.\n"
              "- Make sure ExamGroup is set."
        )

    # OR-Tools import only here
    try:
        from ortools.sat.python import cp_model
    except Exception:
        raise ImportError(
            "OR-Tools is not installed.\n\n"
            "Install with:\n"
            "python -m pip install pandas openpyxl ortools"
        )

    # ---------------- Build CP-SAT model ----------------
    examgroups = sorted([g for g in enroll_df["ExamGroup"].unique().tolist() if str(g).strip() != ""])
    programs = sorted(enroll_df["Program"].unique().tolist())

    cal_df = cal_df.sort_values(["DateN", "StartMin", "SlotID"]).reset_index(drop=True)
    cal_df["SlotKey"] = cal_df["DateN"].astype(str) + " | " + cal_df["SlotID"].astype(str)
    slot_keys = cal_df["SlotKey"].tolist()
    T = len(slot_keys)

    slot_date = cal_df["DateN"].tolist()
    slot_slotid = cal_df["SlotID"].astype(str).tolist()
    slot_start = cal_df["StartMin"].tolist()
    slot_end = cal_df["EndMin"].tolist()
    slot_dur = cal_df["SlotDurationMin"].tolist()

    unique_days = sorted(cal_df["DateN"].unique())
    day_index = {d: i for i, d in enumerate(unique_days)}
    slot_day = [day_index[d] for d in slot_date]
    D = len(unique_days)

    cap_map = dict(zip(cap_df["SlotKey"], cap_df["CapacityStudents"]))
    capacities = [int(cap_map.get(k, 10**9)) for k in slot_keys]

    g_students = enroll_df.groupby("ExamGroup")["StudentID"].nunique().to_dict()
    g_duration = enroll_df.groupby("ExamGroup")["DurationMin"].max().to_dict()
    g_coursecodes = enroll_df.groupby("ExamGroup")["CourseCode"].apply(lambda s: ", ".join(sorted(set(map(str, s))))).to_dict()
    g_coursenames = enroll_df.groupby("ExamGroup")["CourseName"].apply(lambda s: ", ".join(sorted(set(map(str, s))))).to_dict()

    feasible_slots_for_g = {}
    for g in examgroups:
        dur = int(g_duration.get(g, 120))
        feasible = [t for t in range(T) if int(slot_dur[t]) >= dur]
        if not feasible:
            raise ValueError(
                f"ExamGroup '{g}' duration {dur} min cannot fit in any slot.\n"
                "Fix: add longer slots to exam_calendar.xlsx or reduce duration."
            )
        feasible_slots_for_g[g] = feasible

    student_groups = enroll_df.groupby("StudentID")["ExamGroup"].apply(lambda s: sorted(set([x for x in s if str(x).strip() != ""]))).to_dict()

    # Fixed mapping
    fixed_map = {}
    if fixed_df is not None and not fixed_df.empty:
        sk_to_t = {k: i for i, k in enumerate(slot_keys)}
        for _, r in fixed_df.iterrows():
            eg = normalize_str(r.get("ExamGroup"))
            sk = normalize_str(r.get("SlotKey"))
            if not eg and not sk:
                continue
            if eg not in set(examgroups):
                raise ValueError(f"FixedAssignments: ExamGroup not present in enrollments: {eg}")
            if sk not in sk_to_t:
                raise ValueError(f"FixedAssignments: slot not found in calendar: {sk}")
            tfix = sk_to_t[sk]
            if tfix not in feasible_slots_for_g[eg]:
                raise ValueError(f"FixedAssignments: slot duration too short for ExamGroup: {eg}")
            fixed_map[eg] = tfix

    # Weights
    w_capacity, w_rest, w_spread = 50, 30, 5
    if balance_df is not None and not balance_df.empty:
        row0 = balance_df.iloc[0].to_dict()
        w_capacity = int(row0.get("WeightCapacity", w_capacity))
        w_rest = int(row0.get("WeightRestViolation", w_rest))
        w_spread = int(row0.get("WeightSpread", w_spread))

    model = cp_model.CpModel()

    x = {}
    for g in examgroups:
        for t in feasible_slots_for_g[g]:
            x[(g, t)] = model.NewBoolVar(f"x_{g}_{t}")

    for g in examgroups:
        model.Add(sum(x[(g, t)] for t in feasible_slots_for_g[g]) == 1)

    for g, tfix in fixed_map.items():
        model.Add(x[(g, tfix)] == 1)

    # day vars
    day_var = {}
    for g in examgroups:
        dv = model.NewIntVar(0, D - 1, f"day_{g}")
        model.Add(dv == sum(int(slot_day[t]) * x[(g, t)] for t in feasible_slots_for_g[g]))
        day_var[g] = dv

    # Hard: no 2 exams same slot per student
    for sid, glist in student_groups.items():
        if len(glist) < 2:
            continue
        for t in range(T):
            terms = [x[(g, t)] for g in glist if (g, t) in x]
            if len(terms) >= 2:
                model.Add(sum(terms) <= 1)

    # Soft: rest day violations (gap >= rest_days+1 desired)
    rest_violations = []
    max_day = D - 1
    rd = max(0, int(rest_days))

    for sid, glist in student_groups.items():
        if len(glist) < 2:
            continue
        for i in range(len(glist)):
            for j in range(i + 1, len(glist)):
                a, b = glist[i], glist[j]
                diff = model.NewIntVar(0, max_day, f"diff_{sid}_{i}_{j}")
                model.AddAbsEquality(diff, day_var[a] - day_var[b])

                viol = model.NewBoolVar(f"restviol_{sid}_{i}_{j}")
                model.Add(diff <= rd).OnlyEnforceIf(viol)
                model.Add(diff >= rd + 1).OnlyEnforceIf(viol.Not())
                rest_violations.append(viol)

    # Soft: capacity overage
    used_students_slot = []
    over_vars = []
    for t in range(T):
        used = model.NewIntVar(0, 10**9, f"used_{t}")
        terms = [int(g_students[g]) * x[(g, t)] for g in examgroups if (g, t) in x]
        model.Add(used == (sum(terms) if terms else 0))

        cap = int(capacities[t])
        over = model.NewIntVar(0, 10**9, f"over_{t}")
        model.Add(over >= used - cap)
        model.Add(over >= 0)

        used_students_slot.append(used)
        over_vars.append(over)

    # Soft: balance across days
    day_load = []
    for d in range(D):
        load = model.NewIntVar(0, 10**9, f"dayload_{d}")
        terms = []
        for t in range(T):
            if int(slot_day[t]) != d:
                continue
            for g in examgroups:
                if (g, t) in x:
                    terms.append(int(g_students[g]) * x[(g, t)])
        model.Add(load == (sum(terms) if terms else 0))
        day_load.append(load)

    max_load = model.NewIntVar(0, 10**9, "max_dayload")
    min_load = model.NewIntVar(0, 10**9, "min_dayload")
    model.AddMaxEquality(max_load, day_load)
    model.AddMinEquality(min_load, day_load)
    spread = model.NewIntVar(0, 10**9, "spread_dayload")
    model.Add(spread == max_load - min_load)

    obj = []
    if rest_violations:
        obj.append(w_rest * sum(rest_violations))
    if over_vars:
        obj.append(w_capacity * sum(over_vars))
    obj.append(w_spread * spread)
    model.Minimize(sum(obj))

    # ---------------- Solve ----------------
    solver = cp_model.CpSolver()
    solver.parameters.max_time_in_seconds = float(time_limit_sec)
    solver.parameters.num_search_workers = int(workers)

    status = solver.Solve(model)
    status_name = solver.StatusName(status)

    if status not in (cp_model.OPTIMAL, cp_model.FEASIBLE):
        raise RuntimeError(
            f"No feasible solution. Status={status_name}\n\n"
            "Try:\n"
            "- Add more days/slots\n"
            "- Increase capacities\n"
            "- Reduce fixed constraints\n"
            "- Add longer slots for long exams"
        )

    # assignment
    assign = {}
    for g in examgroups:
        chosen = None
        for t in feasible_slots_for_g[g]:
            if solver.Value(x[(g, t)]) == 1:
                chosen = t
                break
        assign[g] = chosen if chosen is not None else feasible_slots_for_g[g][0]

    # MasterSchedule
    master_rows = []
    for g, t in assign.items():
        master_rows.append({
            "ExamGroup": g,
            "CourseCodes": g_coursecodes.get(g, ""),
            "CourseNames": g_coursenames.get(g, ""),
            "TotalStudents": int(g_students.get(g, 0)),
            "Date": mmdd_str(slot_date[t]),
            "DayIndex": int(slot_day[t]),
            "SlotID": slot_slotid[t],
            "Start": fmt_hhmm(slot_start[t]),
            "End": fmt_hhmm(slot_end[t]),
            "SlotDurationMin": int(slot_dur[t]),
            "SlotKey": slot_keys[t],
        })
    master_df = pd.DataFrame(master_rows).sort_values(["DayIndex", "Start", "SlotID", "ExamGroup"])

    # CapacityReport
    cap_rows = []
    for t in range(T):
        used = int(solver.Value(used_students_slot[t]))
        cap = int(capacities[t])
        over = max(0, used - cap) if cap < 10**9 else 0
        cap_rows.append({
            "Date": mmdd_str(slot_date[t]),
            "SlotID": slot_slotid[t],
            "Start": fmt_hhmm(slot_start[t]),
            "End": fmt_hhmm(slot_end[t]),
            "SlotDurationMin": int(slot_dur[t]),
            "CapacityStudents": cap if cap < 10**9 else None,
            "UsedStudents": used,
            "Over": over,
            "SlotKey": slot_keys[t],
        })
    cap_report_df = pd.DataFrame(cap_rows).sort_values(["Date", "Start", "SlotID"])

    # StudentRestViolations
    viol_rows = []
    for sid, glist in student_groups.items():
        if len(glist) < 2:
            continue
        for i in range(len(glist)):
            for j in range(i + 1, len(glist)):
                a, b = glist[i], glist[j]
                gap = abs(int(solver.Value(day_var[a])) - int(solver.Value(day_var[b])))
                if gap <= rd:
                    ta = assign[a]
                    tb = assign[b]
                    prog = enroll_df.loc[enroll_df["StudentID"] == sid, "Program"].iloc[0]
                    viol_rows.append({
                        "StudentID": sid,
                        "Program": prog,
                        "ExamA": a,
                        "DateA": mmdd_str(slot_date[ta]),
                        "SlotA": slot_slotid[ta],
                        "ExamB": b,
                        "DateB": mmdd_str(slot_date[tb]),
                        "SlotB": slot_slotid[tb],
                        "GapDays": gap,
                    })
    rest_viol_df = pd.DataFrame(viol_rows).sort_values(["Program", "StudentID", "GapDays"]) if viol_rows else pd.DataFrame()

    # Program sheets
    prog_group_counts = enroll_df.groupby(["Program", "ExamGroup"])["StudentID"].nunique().reset_index()
    prog_sheets = {}
    for prog in programs:
        egs = prog_group_counts.loc[prog_group_counts["Program"] == prog, "ExamGroup"].tolist()
        tmp = master_df[master_df["ExamGroup"].isin(egs)].copy()
        prog_counts_map = prog_group_counts.loc[prog_group_counts["Program"] == prog].set_index("ExamGroup")["StudentID"].to_dict()
        tmp["ProgramStudents"] = tmp["ExamGroup"].map(lambda gg: int(prog_counts_map.get(gg, 0)))
        prog_sheets[prog] = tmp.sort_values(["DayIndex", "Start", "SlotID", "ExamGroup"])

    # Summary
    summary = {
        "SolverStatus": status_name,
        "ObjectiveValue": float(solver.ObjectiveValue()),
        "TotalStudents": int(enroll_df["StudentID"].nunique()),
        "TotalPrograms": int(len(programs)),
        "TotalExamGroups": int(len(examgroups)),
        "TotalSlots": int(T),
        "UniqueDays": int(D),
        "RestDaysSoft": int(rd),
        "WeightRestViolation": int(w_rest),
        "WeightCapacity": int(w_capacity),
        "WeightSpread": int(w_spread),
        "SlotsOverCapacity": int((cap_report_df["Over"] > 0).sum()) if not cap_report_df.empty else 0,
        "RestViolationsPairs": int(len(rest_viol_df)) if not rest_viol_df.empty else 0,
    }
    summary_df = pd.DataFrame([summary])

    # Save output
    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
        master_df.to_excel(writer, index=False, sheet_name="MasterSchedule")
        cap_report_df.to_excel(writer, index=False, sheet_name="CapacityReport")
        summary_df.to_excel(writer, index=False, sheet_name="Summary")

        if rest_viol_df is not None and not rest_viol_df.empty:
            rest_viol_df.to_excel(writer, index=False, sheet_name="StudentRestViolations")

        for prog, dfp in prog_sheets.items():
            dfp.to_excel(writer, index=False, sheet_name=safe_sheet_name(f"Program_{prog}"))

    return master_df, prog_sheets, cap_report_df, rest_viol_df, summary_df

